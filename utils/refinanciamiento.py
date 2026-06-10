import json
import re
from datetime import date, datetime
from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle
)


COLUMNAS_MANUALES = [
    "INCLUIR",
    "FACT",
    "VTA",
    "SALDO",
    "ABONO",
    "EN COBRO",
    "QNAS TOMADAS A CUENTA"
]

COLUMNAS_CALCULADAS = [
    "PAGADO",
    "ABONO DE QUINCENAS CONS",
    "SALDO PENDIENTE",
    "PORCENTAJE PAGADO",
    "REFINANCIAMIENTO",
    "PUEDE REFINANCIAR",
    "ESTADO"
]

PLAZOS_REFINANCIAMIENTO = [72, 60, 46, 34]
OUTPUT_REFINANCIAMIENTO_DIR = Path("REFINANCIAMIENTOS")


def to_number(valor) -> float:
    if valor is None or pd.isna(valor):
        return 0.0

    if isinstance(valor, str):
        valor = (
            valor.replace("$", "")
            .replace(",", "")
            .replace("%", "")
            .strip()
        )

    try:
        return float(valor)
    except (TypeError, ValueError):
        return 0.0


convertir_numero = to_number


def slug_folder_name(valor, reemplazar_espacios: bool = False) -> str:
    texto = str(valor or "").strip()
    texto = re.sub(r'[\\/:*?"<>|]', "", texto)
    texto = " ".join(texto.split())

    if reemplazar_espacios:
        texto = texto.replace(" ", "_")

    return texto


def construir_carpeta_refinanciamiento(
    base_dir,
    semana: int,
    vendedor: str,
    cliente: str
) -> Path:
    semana = max(1, min(int(semana), 53))
    vendedor_limpio = (
        slug_folder_name(vendedor)
        or "VENDEDOR_SIN_SELECCIONAR"
    )
    cliente_limpio = (
        slug_folder_name(cliente)
        or "CLIENTE_SIN_NOMBRE"
    )

    return (
        Path(base_dir)
        / f"SEM_{semana:02d}"
        / vendedor_limpio
        / cliente_limpio
    )


def nombre_archivo_refinanciamiento(
    cliente: str = "",
    semana: int = 1,
    extension: str = "xlsx"
) -> str:
    extension = str(extension or "xlsx").lower().lstrip(".")
    return f"refinanciamiento.{extension}"


def guardar_excel_refinanciamiento(
    contenido: bytes,
    base_dir,
    semana: int,
    vendedor: str,
    cliente: str
) -> Path:
    carpeta = construir_carpeta_refinanciamiento(
        base_dir=base_dir,
        semana=semana,
        vendedor=vendedor,
        cliente=cliente
    )
    carpeta.mkdir(parents=True, exist_ok=True)
    ruta = carpeta / nombre_archivo_refinanciamiento(
        cliente,
        semana,
        "xlsx"
    )
    ruta.write_bytes(contenido)
    return ruta


def guardar_archivos_refinanciamiento(
    archivos: dict,
    base_dir,
    semana: int,
    vendedor: str,
    cliente: str
) -> dict:
    carpeta = construir_carpeta_refinanciamiento(
        base_dir=base_dir,
        semana=semana,
        vendedor=vendedor,
        cliente=cliente
    )
    carpeta.mkdir(parents=True, exist_ok=True)
    rutas = {}

    for extension, contenido in archivos.items():
        extension_limpia = str(extension).lower().lstrip(".")
        ruta = carpeta / nombre_archivo_refinanciamiento(
            cliente,
            semana,
            extension_limpia
        )
        ruta.write_bytes(contenido)
        rutas[extension_limpia] = ruta

    return rutas


def extraer_fecha_edad_desde_rfc(
    rfc: str,
    forzar_1900: bool = True,
    fecha_referencia: date = None
) -> dict:
    fecha_referencia = fecha_referencia or date.today()
    rfc = str(rfc or "").upper().strip()

    for coincidencia in re.finditer(r"\d{6}", rfc):
        bloque = coincidencia.group(0)
        anio_corto = int(bloque[:2])
        mes = int(bloque[2:4])
        dia = int(bloque[4:6])

        if forzar_1900:
            anio = 1900 + anio_corto
        else:
            anio_actual_corto = fecha_referencia.year % 100
            siglo = 2000 if anio_corto <= anio_actual_corto else 1900
            anio = siglo + anio_corto

        try:
            fecha_nacimiento = date(anio, mes, dia)
        except ValueError:
            continue

        if fecha_nacimiento > fecha_referencia:
            continue

        edad = fecha_referencia.year - fecha_nacimiento.year
        if (
            fecha_referencia.month,
            fecha_referencia.day
        ) < (
            fecha_nacimiento.month,
            fecha_nacimiento.day
        ):
            edad -= 1

        return {
            "fecha_nacimiento": fecha_nacimiento,
            "edad": edad
        }

    return {
        "fecha_nacimiento": None,
        "edad": None
    }


def dataframe_facturas_vacio(filas: int = 5) -> pd.DataFrame:
    return pd.DataFrame([
        {
            "INCLUIR": False,
            "FACT": "",
            "VTA": 0.0,
            "SALDO": 0.0,
            "ABONO": 0.0,
            "EN COBRO": "",
            "QNAS TOMADAS A CUENTA": ""
        }
        for _ in range(filas)
    ])


def preparar_facturas_desde_bd(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return dataframe_facturas_vacio(0)

    facturas = df.rename(columns={
        "fact": "FACT",
        "vta": "VTA",
        "saldo": "SALDO",
        "venta_id": "VENTA_ID"
    }).copy()
    facturas["ABONO"] = 0.0
    facturas["EN COBRO"] = ""
    facturas["QNAS TOMADAS A CUENTA"] = 0
    facturas["INCLUIR"] = True

    calculadas = calcular_facturas_refinanciamiento(facturas)
    calculadas["INCLUIR"] = calculadas["PUEDE REFINANCIAR"].eq("SI")
    return calculadas


def calcular_facturas_refinanciamiento(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=COLUMNAS_MANUALES + COLUMNAS_CALCULADAS)

    resultado = df.copy()

    for columna in COLUMNAS_MANUALES:
        if columna not in resultado:
            if columna == "INCLUIR":
                resultado[columna] = pd.NA
            else:
                resultado[columna] = (
                    "" if columna in ["FACT", "EN COBRO"] else 0.0
                )

    columnas_numericas = [
        "VTA",
        "SALDO",
        "QNAS TOMADAS A CUENTA",
        "ABONO"
    ]
    for columna in columnas_numericas:
        resultado[columna] = resultado[columna].map(convertir_numero)

    resultado["PAGADO"] = 0.0
    tiene_venta = resultado["VTA"] > 0
    resultado.loc[tiene_venta, "PAGADO"] = (
        resultado.loc[tiene_venta, "VTA"]
        - resultado.loc[tiene_venta, "SALDO"]
    ).round(2)
    resultado["ABONO DE QUINCENAS CONS"] = (
        resultado["QNAS TOMADAS A CUENTA"] * resultado["ABONO"]
    ).round(2)
    resultado["SALDO PENDIENTE"] = (
        resultado["SALDO"] - resultado["ABONO DE QUINCENAS CONS"]
    ).round(2)
    resultado["PORCENTAJE PAGADO"] = 0.0
    resultado.loc[tiene_venta, "PORCENTAJE PAGADO"] = (
        resultado.loc[tiene_venta, "PAGADO"]
        / resultado.loc[tiene_venta, "VTA"]
    )
    resultado["PORCENTAJE PAGADO"] = resultado[
        "PORCENTAJE PAGADO"
    ].round(6)
    resultado["REFINANCIAMIENTO"] = resultado["PAGADO"].round(2)
    fila_activa = (
        resultado["FACT"].astype(str).str.strip().ne("")
        | resultado["EN COBRO"].astype(str).str.strip().ne("")
        | resultado[columnas_numericas].abs().sum(axis=1).gt(0)
    )
    resultado["PUEDE REFINANCIAR"] = ""
    resultado.loc[fila_activa, "PUEDE REFINANCIAR"] = resultado.loc[
        fila_activa,
        "PORCENTAJE PAGADO"
    ].map(lambda porcentaje: "SI" if porcentaje >= 0.40 else "NO")
    resultado["ESTADO"] = ""
    resultado.loc[fila_activa, "ESTADO"] = resultado.loc[
        fila_activa,
        "PUEDE REFINANCIAR"
    ].map({"SI": "APTA", "NO": "NO APTA"})

    incluir_original = resultado["INCLUIR"]
    incluir_default = resultado["PUEDE REFINANCIAR"].eq("SI")
    resultado["INCLUIR"] = incluir_original.where(
        incluir_original.notna(),
        incluir_default
    ).map(lambda valor: bool(valor) if not pd.isna(valor) else False)

    columnas_salida = [
        "INCLUIR",
        "FACT",
        "VTA",
        "PAGADO",
        "SALDO",
        "QNAS TOMADAS A CUENTA",
        "ABONO",
        "ABONO DE QUINCENAS CONS",
        "SALDO PENDIENTE",
        "EN COBRO",
        "PORCENTAJE PAGADO",
        "REFINANCIAMIENTO",
        "PUEDE REFINANCIAR",
        "ESTADO"
    ]
    return resultado[columnas_salida]


def filtrar_facturas_capturadas(df: pd.DataFrame) -> pd.DataFrame:
    facturas = calcular_facturas_refinanciamiento(df)

    if facturas.empty:
        return facturas

    columnas_numericas = [
        "VTA",
        "SALDO",
        "QNAS TOMADAS A CUENTA",
        "ABONO"
    ]
    fila_activa = (
        facturas["FACT"].astype(str).str.strip().ne("")
        | facturas["EN COBRO"].astype(str).str.strip().ne("")
        | facturas[columnas_numericas].abs().sum(axis=1).gt(0)
    )
    return facturas.loc[fila_activa].reset_index(drop=True)


def calcular_resumen_refinanciamiento(
    df: pd.DataFrame,
    aumento_descuento: float
) -> dict:
    facturas = calcular_facturas_refinanciamiento(df)
    facturas = facturas[facturas["INCLUIR"]].copy()
    aumento_descuento = convertir_numero(aumento_descuento)

    def total(columna: str) -> float:
        if facturas.empty:
            return 0.0
        return round(float(facturas[columna].sum()), 2)

    abono_ref = total("ABONO")
    abono_antes = total("ABONO")
    total_abono_nuevo = round(abono_ref + aumento_descuento, 2)
    total_saldo_pendiente = total("SALDO PENDIENTE")

    simulacion = {}
    for plazo in PLAZOS_REFINANCIAMIENTO:
        total_saldo_nuevo = round(total_abono_nuevo * plazo, 2)
        venta_posible = round(
            total_saldo_nuevo - total_saldo_pendiente,
            2
        )
        simulacion[plazo] = {
            "SALDO PENDIENTE": total_saldo_pendiente,
            "VENTA POSIBLE": venta_posible,
            "TOTAL SALDO NUEVO": total_saldo_nuevo,
            "DESCUENTO NUEVO": total_abono_nuevo,
            "TOTAL ADEUDO CLIENTE": total_saldo_nuevo
        }

    return {
        "total_vta": total("VTA"),
        "total_pagado": total("PAGADO"),
        "total_saldo": total("SALDO"),
        "total_abono": abono_antes,
        "total_abono_quincenas": total("ABONO DE QUINCENAS CONS"),
        "total_saldo_pendiente": total_saldo_pendiente,
        "total_refinanciamiento": total("REFINANCIAMIENTO"),
        "abono_ref": abono_ref,
        "abono_antes": abono_antes,
        "aumento_descuento": round(aumento_descuento, 2),
        "total_abono_nuevo": total_abono_nuevo,
        "simulacion": simulacion
    }


def dataframe_simulacion(resumen: dict) -> pd.DataFrame:
    filas = [
        "SALDO PENDIENTE",
        "VENTA POSIBLE",
        "TOTAL SALDO NUEVO",
        "DESCUENTO NUEVO",
        "TOTAL ADEUDO CLIENTE"
    ]

    return pd.DataFrame({
        plazo: [
            resumen["simulacion"][plazo][fila]
            for fila in filas
        ]
        for plazo in PLAZOS_REFINANCIAMIENTO
    }, index=filas)


def generar_excel_refinanciamiento(
    facturas: pd.DataFrame,
    resumen: dict,
    datos_cliente: dict
) -> bytes:
    facturas = filtrar_facturas_capturadas(facturas)
    wb = Workbook()
    ws_facturas = wb.active
    ws_facturas.title = "Facturas"

    encabezado_fill = PatternFill("solid", fgColor="1F4E78")
    encabezado_font = Font(color="FFFFFF", bold=True)
    subtitulo_fill = PatternFill("solid", fgColor="D9EAF7")
    moneda = '$ #,##0.00;-$ #,##0.00;$ -'
    porcentaje = "0%"

    columnas_facturas = [
        "INCLUIR",
        "FACT",
        "VTA",
        "PAGADO",
        "SALDO",
        "QNAS TOMADAS A CUENTA",
        "ABONO",
        "ABONO DE QUINCENAS CONS",
        "SALDO PENDIENTE",
        "EN COBRO",
        "PORCENTAJE PAGADO",
        "REFINANCIAMIENTO",
        "PUEDE REFINANCIAR",
        "ESTADO"
    ]
    ws_facturas.append(columnas_facturas)

    for celda in ws_facturas[1]:
        celda.fill = encabezado_fill
        celda.font = encabezado_font
        celda.alignment = Alignment(horizontal="center", wrap_text=True)

    for _, fila in facturas.iterrows():
        ws_facturas.append([fila[columna] for columna in columnas_facturas])

    columnas_moneda = {
        "VTA",
        "PAGADO",
        "SALDO",
        "ABONO",
        "ABONO DE QUINCENAS CONS",
        "SALDO PENDIENTE",
        "REFINANCIAMIENTO"
    }
    for indice, columna in enumerate(columnas_facturas, start=1):
        ancho = min(max(len(columna) + 3, 13), 27)
        ws_facturas.column_dimensions[get_column_letter(indice)].width = ancho
        for celda in ws_facturas[get_column_letter(indice)][1:]:
            if columna in columnas_moneda:
                celda.number_format = moneda
            elif columna == "PORCENTAJE PAGADO":
                celda.number_format = porcentaje

    ws_facturas.freeze_panes = "A2"
    ws_facturas.auto_filter.ref = ws_facturas.dimensions

    ws_resumen = wb.create_sheet("Resumen Refinanciamiento")
    ws_resumen["A1"] = "RESUMEN DE REFINANCIAMIENTO"
    ws_resumen["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws_resumen["A1"].fill = encabezado_fill
    ws_resumen.merge_cells("A1:E1")
    ws_resumen["A1"].alignment = Alignment(horizontal="center")

    campos = [
        ("Semana", datos_cliente.get("semana")),
        ("Vendedor", datos_cliente.get("vendedor", "")),
        ("Fecha", datos_cliente.get("fecha")),
        ("Cliente", datos_cliente.get("cliente", "")),
        ("RFC/NAC", datos_cliente.get("rfc_nac", "")),
        ("Fecha nacimiento", datos_cliente.get("fecha_nacimiento")),
        ("Edad", datos_cliente.get("edad")),
        ("Quinquenio", datos_cliente.get("quinquenio", "")),
        ("ABONO REF", resumen["abono_ref"]),
        ("ABONO ANTES", resumen["abono_antes"]),
        ("LIQUIDEZ / AUMENTO EN DESCUENTO", resumen["aumento_descuento"]),
        ("TOTAL ABONO NUEVO", resumen["total_abono_nuevo"]),
        ("total_vta", resumen["total_vta"]),
        ("total_pagado", resumen["total_pagado"]),
        ("total_saldo", resumen["total_saldo"]),
        ("total_abono", resumen["total_abono"]),
        ("total_saldo_pendiente", resumen["total_saldo_pendiente"])
    ]

    for fila, (etiqueta, valor) in enumerate(campos, start=3):
        ws_resumen.cell(fila, 1, etiqueta)
        ws_resumen.cell(fila, 1).font = Font(bold=True)
        ws_resumen.cell(fila, 2, valor)

        if fila >= 11:
            ws_resumen.cell(fila, 2).number_format = moneda

    ws_resumen["B5"].number_format = "dd/mm/yyyy"
    ws_resumen["B8"].number_format = "dd/mm/yyyy"

    inicio_simulacion = 22
    ws_resumen.cell(inicio_simulacion, 1, "PLAZO")
    for columna, plazo in enumerate(PLAZOS_REFINANCIAMIENTO, start=2):
        ws_resumen.cell(inicio_simulacion, columna, plazo)

    filas_simulacion = [
        "SALDO PENDIENTE",
        "VENTA POSIBLE",
        "TOTAL SALDO NUEVO",
        "DESCUENTO NUEVO",
        "TOTAL ADEUDO CLIENTE"
    ]
    for desplazamiento, nombre_fila in enumerate(filas_simulacion, start=1):
        fila_excel = inicio_simulacion + desplazamiento
        ws_resumen.cell(fila_excel, 1, nombre_fila)
        for columna, plazo in enumerate(PLAZOS_REFINANCIAMIENTO, start=2):
            celda = ws_resumen.cell(
                fila_excel,
                columna,
                resumen["simulacion"][plazo][nombre_fila]
            )
            celda.number_format = moneda

    for celda in ws_resumen[inicio_simulacion]:
        celda.fill = encabezado_fill
        celda.font = encabezado_font
        celda.alignment = Alignment(horizontal="center")

    for fila in range(inicio_simulacion + 1, inicio_simulacion + 6):
        ws_resumen.cell(fila, 1).fill = subtitulo_fill
        ws_resumen.cell(fila, 1).font = Font(bold=True)

    for fila in [inicio_simulacion + 2, inicio_simulacion + 4]:
        for columna in range(1, 6):
            celda = ws_resumen.cell(fila, columna)
            celda.fill = PatternFill("solid", fgColor="FFF2CC")
            celda.font = Font(bold=True)

    ws_resumen.column_dimensions["A"].width = 35
    for columna in "BCDE":
        ws_resumen.column_dimensions[columna].width = 18
    ws_resumen.freeze_panes = "A3"

    salida = BytesIO()
    wb.save(salida)
    return salida.getvalue()


def _valor_json(valor):
    if pd.isna(valor):
        return None
    if isinstance(valor, (date, datetime)):
        return valor.isoformat()
    if hasattr(valor, "item"):
        return valor.item()
    return valor


def generar_json_refinanciamiento(
    facturas: pd.DataFrame,
    resumen: dict,
    datos_cliente: dict
) -> bytes:
    facturas_calculadas = filtrar_facturas_capturadas(facturas)
    payload = {
        "datos_cliente": {
            clave: _valor_json(valor)
            for clave, valor in datos_cliente.items()
        },
        "facturas": [
            {
                clave: _valor_json(valor)
                for clave, valor in fila.items()
            }
            for fila in facturas_calculadas.to_dict(orient="records")
        ],
        "resumen": {
            clave: valor
            for clave, valor in resumen.items()
            if clave != "simulacion"
        },
        "simulacion": resumen["simulacion"]
    }
    return json.dumps(
        payload,
        ensure_ascii=False,
        indent=2
    ).encode("utf-8")


def generar_pdf_refinanciamiento(
    facturas: pd.DataFrame,
    resumen: dict,
    datos_cliente: dict
) -> bytes:
    salida = BytesIO()
    documento = SimpleDocTemplate(
        salida,
        pagesize=landscape(letter),
        rightMargin=0.35 * inch,
        leftMargin=0.35 * inch,
        topMargin=0.35 * inch,
        bottomMargin=0.35 * inch
    )
    estilos = getSampleStyleSheet()
    elementos = [
        Paragraph("Resumen de refinanciamiento", estilos["Title"]),
        Paragraph(
            (
                f"Cliente: {datos_cliente.get('cliente', '')} | "
                f"RFC: {datos_cliente.get('rfc_nac', '')} | "
                f"Vendedor: {datos_cliente.get('vendedor', '')} | "
                f"Semana: {datos_cliente.get('semana', '')}"
            ),
            estilos["Normal"]
        ),
        Spacer(1, 10)
    ]

    facturas_calculadas = filtrar_facturas_capturadas(facturas)
    encabezados = [
        "INCLUIR", "FACT", "VTA", "PAGADO", "SALDO", "QNAS",
        "ABONO", "SALDO PEND.", "% PAGADO", "ESTADO"
    ]
    filas = [encabezados]
    for _, factura in facturas_calculadas.iterrows():
        filas.append([
            "SI" if factura["INCLUIR"] else "NO",
            str(factura["FACT"]),
            f"${factura['VTA']:,.2f}",
            f"${factura['PAGADO']:,.2f}",
            f"${factura['SALDO']:,.2f}",
            f"{factura['QNAS TOMADAS A CUENTA']:,.0f}",
            f"${factura['ABONO']:,.2f}",
            f"${factura['SALDO PENDIENTE']:,.2f}",
            f"{factura['PORCENTAJE PAGADO']:.0%}",
            factura["ESTADO"]
        ])

    tabla_facturas = Table(filas, repeatRows=1)
    tabla_facturas.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
        ("ALIGN", (2, 1), (-2, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE")
    ]))
    elementos.extend([tabla_facturas, Spacer(1, 12)])

    resumen_filas = [
        ["Concepto", "Importe"],
        ["Total VTA", f"${resumen['total_vta']:,.2f}"],
        ["Total pagado", f"${resumen['total_pagado']:,.2f}"],
        ["Saldo pendiente", f"${resumen['total_saldo_pendiente']:,.2f}"],
        ["Descuento nuevo", f"${resumen['total_abono_nuevo']:,.2f}"]
    ]
    tabla_resumen = Table(resumen_filas, colWidths=[2.3 * inch, 1.5 * inch])
    tabla_resumen.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
        ("ALIGN", (1, 1), (1, -1), "RIGHT")
    ]))
    elementos.append(tabla_resumen)
    documento.build(elementos)
    return salida.getvalue()
