import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader

from services.calculadora import calcular_revision_talon
from services.generador_excel import generar_excel_revision
from services.generador_pdf import generar_pdf_revision


class CuentasTerminadasTest(unittest.TestCase):
    def setUp(self):
        self.codigos = {
            "E4": {"descripcion": "Ingreso", "importe": 10000.0}
        }
        self.cuentas = [
            {
                "qna_termina": "16-2026",
                "saldo_liberado": 1500.0,
                "sumar_a_liquidez": True,
                "observacion": "Cuenta liquidada"
            },
            {
                "qna_termina": "17-2026",
                "saldo_liberado": 800.0,
                "sumar_a_liquidez": False,
                "observacion": ""
            }
        ]

    def test_calculo_separa_saldos_y_solo_suma_cuentas_activas(self):
        revision = calcular_revision_talon(
            codigos_extraidos=self.codigos,
            descuentos_talon=4000,
            abono_extra=500,
            programado=1000,
            cuentas_terminadas=self.cuentas
        )

        self.assertEqual(revision["liquidez_talon"], 6000)
        self.assertEqual(revision["abono_extra"], 500)
        self.assertEqual(revision["total_saldo_liberado"], 1500)
        self.assertEqual(revision["total_solo_observado"], 800)
        self.assertEqual(revision["programado"], 1000)
        self.assertEqual(revision["liquidez_antes_liberacion"], 5500)
        self.assertEqual(revision["liquidez_final"], 7000)

    def test_sin_cuentas_conserva_el_calculo_anterior(self):
        revision = calcular_revision_talon(
            codigos_extraidos=self.codigos,
            descuentos_talon=4000,
            abono_extra=500,
            programado=1000
        )

        self.assertEqual(revision["total_saldo_liberado"], 0)
        self.assertEqual(revision["total_solo_observado"], 0)
        self.assertEqual(revision["liquidez_final"], 5500)

    def test_cuenta_solo_observada_no_modifica_liquidez(self):
        revision = calcular_revision_talon(
            codigos_extraidos=self.codigos,
            descuentos_talon=4000,
            abono_extra=500,
            programado=1000,
            cuentas_terminadas=[{
                "qna_termina": "18-2026",
                "saldo_liberado": 2200,
                "sumar_a_liquidez": False,
                "observacion": "Pendiente de confirmar"
            }]
        )

        self.assertEqual(revision["total_saldo_liberado"], 0)
        self.assertEqual(revision["total_solo_observado"], 2200)
        self.assertEqual(revision["liquidez_final"], 5500)

    def test_una_cuenta_sumada_incrementa_liquidez(self):
        revision = calcular_revision_talon(
            codigos_extraidos=self.codigos,
            descuentos_talon=11000,
            cuentas_terminadas=[{
                "qna_termina": "16-2026",
                "saldo_liberado": 1500,
                "sumar_a_liquidez": True,
                "observacion": ""
            }]
        )

        self.assertEqual(revision["liquidez_talon"], -1000)
        self.assertEqual(revision["total_saldo_liberado"], 1500)
        self.assertEqual(revision["liquidez_final"], 500)

    def test_excel_agrega_resumen_y_hoja_de_detalle(self):
        revision = calcular_revision_talon(
            codigos_extraidos=self.codigos,
            descuentos_talon=4000,
            abono_extra=500,
            programado=1000,
            cuentas_terminadas=self.cuentas
        )
        datos = {
            "nombre": "CLIENTE PRUEBA",
            "rfc": "AAAA010101AAA",
            "fecha_pago": "09/06/2026"
        }
        template = Path("templates/plantilla_revision_talon.xlsx").resolve()

        with tempfile.TemporaryDirectory() as carpeta:
            ruta = generar_excel_revision(
                datos=datos,
                revision=revision,
                mensaje_vendedor="Mensaje de prueba",
                promotor="Victor Vega",
                qna="12-2026",
                ruta_template=str(template),
                carpeta_salida=carpeta
            )
            workbook = load_workbook(ruta, data_only=False)
            principal = workbook["Hoja1"]
            detalle = workbook["CUENTAS_TERMINADAS"]

            self.assertEqual(principal["A41"].value, "LIQUIDEZ DEL TALÓN")
            self.assertEqual(principal["D41"].value, 6000)
            self.assertEqual(principal["D42"].value, 500)
            self.assertEqual(principal["D43"].value, 1500)
            self.assertEqual(principal["D44"].value, 800)
            self.assertEqual(principal["D45"].value, 1000)
            self.assertIn("16-2026", principal["A49"].value)
            self.assertIn("17-2026", principal["A53"].value)
            self.assertEqual(principal["D47"].value, 7000)

            self.assertEqual(detalle.max_row, 3)
            self.assertEqual(detalle["F2"].value, "16-2026")
            self.assertEqual(detalle["G2"].value, 1500)
            self.assertEqual(detalle["H2"].value, "Sí")
            self.assertEqual(detalle["F3"].value, "17-2026")
            self.assertEqual(detalle["H3"].value, "No")

    def test_pdf_incluye_resumen_y_cuentas_terminadas(self):
        revision = calcular_revision_talon(
            codigos_extraidos=self.codigos,
            descuentos_talon=4000,
            abono_extra=500,
            programado=1000,
            cuentas_terminadas=self.cuentas
        )
        datos = {
            "nombre": "CLIENTE PRUEBA",
            "rfc": "AAAA010101AAA",
            "fecha_pago": "09/06/2026"
        }

        with tempfile.TemporaryDirectory() as carpeta:
            ruta = generar_pdf_revision(
                datos=datos,
                revision=revision,
                mensaje_vendedor="Mensaje de prueba",
                promotor="Victor Vega",
                qna="12-2026",
                carpeta_salida=carpeta
            )
            texto = "\n".join(
                pagina.extract_text() or ""
                for pagina in PdfReader(ruta).pages
            )

            self.assertIn("Liquidez del talón", texto)
            self.assertIn("16-2026", texto)
            self.assertIn("17-2026", texto)
            self.assertIn("Mensaje de prueba", texto)


if __name__ == "__main__":
    unittest.main()
