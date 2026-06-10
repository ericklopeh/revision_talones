import unittest
import tempfile
import json
from datetime import date
from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from utils.refinanciamiento import (
    COLUMNAS_MANUALES,
    actualizar_inclusiones_por_sale_id,
    aplicar_inclusiones_por_sale_id,
    calcular_facturas_refinanciamiento,
    calcular_resumen_refinanciamiento,
    construir_carpeta_refinanciamiento,
    dataframe_facturas_vacio,
    extraer_fecha_edad_desde_rfc,
    generar_excel_refinanciamiento,
    generar_json_refinanciamiento,
    generar_pdf_refinanciamiento,
    guardar_archivos_refinanciamiento,
    guardar_excel_refinanciamiento,
    nombre_archivo_refinanciamiento,
    preparar_facturas_desde_bd,
    slug_folder_name,
    to_number
)


class RefinanciamientoTest(unittest.TestCase):
    def setUp(self):
        self.facturas = pd.DataFrame([
            {
                "FACT": "12546",
                "VTA": 58564.00,
                "SALDO": 16446.80,
                "QNAS TOMADAS A CUENTA": 0,
                "ABONO": 813.39,
                "EN COBRO": ""
            },
            {
                "FACT": "12547",
                "VTA": 28693.43,
                "SALDO": 8368.91,
                "QNAS TOMADAS A CUENTA": 0,
                "ABONO": 398.52,
                "EN COBRO": ""
            },
            {
                "FACT": "12622",
                "VTA": 29966.96,
                "SALDO": 9988.88,
                "QNAS TOMADAS A CUENTA": 0,
                "ABONO": 416.21,
                "EN COBRO": ""
            },
            {
                "FACT": "12801",
                "VTA": 22251.39,
                "SALDO": 9580.34,
                "QNAS TOMADAS A CUENTA": 0,
                "ABONO": 309.05,
                "EN COBRO": ""
            },
            {
                "FACT": "12719",
                "VTA": 40994.80,
                "SALDO": 14803.78,
                "QNAS TOMADAS A CUENTA": 0,
                "ABONO": 569.37,
                "EN COBRO": ""
            }
        ])

    def test_extrae_fecha_y_edad_desde_rfc(self):
        resultado = extraer_fecha_edad_desde_rfc(
            "J860901UT6",
            fecha_referencia=date(2026, 6, 9)
        )

        self.assertEqual(resultado["fecha_nacimiento"], date(1986, 9, 1))
        self.assertEqual(resultado["edad"], 39)

    def test_rfc_sin_fecha_valida_no_calcula_edad(self):
        resultado = extraer_fecha_edad_desde_rfc(
            "ABC991332XYZ",
            fecha_referencia=date(2026, 6, 9)
        )

        self.assertIsNone(resultado["fecha_nacimiento"])
        self.assertIsNone(resultado["edad"])

    def test_calculos_por_factura_evitan_division_entre_cero(self):
        facturas = calcular_facturas_refinanciamiento(pd.DataFrame([{
            "FACT": "PRUEBA",
            "VTA": 0,
            "SALDO": 500,
            "QNAS TOMADAS A CUENTA": 2,
            "ABONO": 50,
            "EN COBRO": ""
        }]))

        self.assertEqual(
            facturas.loc[0, "ABONO DE QUINCENAS CONS"],
            100
        )
        self.assertEqual(facturas.loc[0, "PAGADO"], 0)
        self.assertEqual(facturas.loc[0, "SALDO PENDIENTE"], 400)
        self.assertEqual(facturas.loc[0, "PORCENTAJE PAGADO"], 0)
        self.assertEqual(facturas.loc[0, "PUEDE REFINANCIAR"], "NO")

    def test_qnas_vacias_o_invalidas_se_toman_como_cero(self):
        facturas = calcular_facturas_refinanciamiento(pd.DataFrame([
            {
                "FACT": "VACIA",
                "VTA": "$ 1,000.00",
                "SALDO": "600.00",
                "ABONO": "100",
                "EN COBRO": "",
                "QNAS TOMADAS A CUENTA": ""
            },
            {
                "FACT": "TEXTO",
                "VTA": 1000,
                "SALDO": 600,
                "ABONO": 100,
                "EN COBRO": "",
                "QNAS TOMADAS A CUENTA": "dato inválido"
            }
        ]))

        self.assertEqual(
            facturas.loc[0, "ABONO DE QUINCENAS CONS"],
            0
        )
        self.assertEqual(
            facturas.loc[1, "ABONO DE QUINCENAS CONS"],
            0
        )
        self.assertEqual(facturas.loc[0, "SALDO PENDIENTE"], 600)
        self.assertEqual(facturas.loc[1, "SALDO PENDIENTE"], 600)

    def test_columnas_manuales_tienen_orden_y_qnas_vacias(self):
        self.assertEqual(
            COLUMNAS_MANUALES,
            [
                "INCLUIR",
                "FACT",
                "VTA",
                "SALDO",
                "ABONO",
                "EN COBRO",
                "QNAS TOMADAS A CUENTA"
            ]
        )
        self.assertEqual(
            dataframe_facturas_vacio(1).loc[0, "QNAS TOMADAS A CUENTA"],
            ""
        )

    def test_conversion_numerica_acepta_formatos_comunes(self):
        self.assertEqual(to_number("$ 58,564.00"), 58564)
        self.assertEqual(to_number("72%"), 72)
        self.assertEqual(to_number(""), 0)

    def test_totales_y_simulacion_coinciden_con_ejemplo(self):
        resumen = calcular_resumen_refinanciamiento(self.facturas, 0)

        self.assertEqual(resumen["total_vta"], 180470.58)
        self.assertEqual(resumen["total_pagado"], 121281.87)
        self.assertEqual(resumen["total_saldo"], 59188.71)
        self.assertEqual(resumen["total_abono"], 2506.54)
        self.assertEqual(resumen["abono_ref"], 2506.54)
        self.assertEqual(resumen["total_saldo_pendiente"], 59188.71)
        self.assertEqual(
            resumen["simulacion"][72]["VENTA POSIBLE"],
            121282.17
        )
        self.assertEqual(
            resumen["simulacion"][34]["TOTAL SALDO NUEVO"],
            85222.36
        )
        self.assertEqual(
            resumen["simulacion"][60]["VENTA POSIBLE"],
            91203.69
        )
        self.assertEqual(
            resumen["simulacion"][46]["TOTAL ADEUDO CLIENTE"],
            115300.84
        )

    def test_umbral_de_refinanciamiento_y_aumento(self):
        facturas = pd.DataFrame([
            {
                "FACT": "SI",
                "VTA": 1000,
                "SALDO": 600,
                "QNAS TOMADAS A CUENTA": 0,
                "ABONO": 100,
                "EN COBRO": ""
            },
            {
                "FACT": "NO",
                "VTA": 1000,
                "SALDO": 601,
                "QNAS TOMADAS A CUENTA": 0,
                "ABONO": 200,
                "EN COBRO": ""
            }
        ])
        resultado = calcular_facturas_refinanciamiento(facturas)
        resumen = calcular_resumen_refinanciamiento(resultado, 50)

        self.assertEqual(resultado.loc[0, "PUEDE REFINANCIAR"], "SI")
        self.assertEqual(resultado.loc[1, "PUEDE REFINANCIAR"], "NO")
        self.assertEqual(resumen["abono_ref"], 100)
        self.assertEqual(resumen["abono_antes"], 100)
        self.assertEqual(resumen["total_abono_nuevo"], 150)

    def test_pagado_manual_se_ignora_y_se_recalcula(self):
        resultado = calcular_facturas_refinanciamiento(pd.DataFrame([{
            "FACT": "12546",
            "VTA": 58564,
            "PAGADO": 1,
            "SALDO": 16446.80,
            "QNAS TOMADAS A CUENTA": 0,
            "ABONO": 813.39,
            "EN COBRO": ""
        }]))

        self.assertEqual(resultado.loc[0, "PAGADO"], 42117.20)
        self.assertAlmostEqual(
            resultado.loc[0, "PORCENTAJE PAGADO"],
            71.92,
            places=2
        )
        self.assertEqual(resultado.loc[0, "REFINANCIAMIENTO"], 42117.20)
        self.assertEqual(resultado.loc[0, "PUEDE REFINANCIAR"], "SI")

    def test_excel_exportado_tiene_dos_hojas(self):
        resumen = calcular_resumen_refinanciamiento(self.facturas, 0)
        contenido = generar_excel_refinanciamiento(
            facturas=self.facturas,
            resumen=resumen,
            datos_cliente={
                "fecha": date(2026, 6, 9),
                "semana": 23,
                "vendedor": "Juan Manuel",
                "cliente": "CLIENTE PRUEBA",
                "rfc_nac": "J860901UT6",
                "fecha_nacimiento": date(1986, 9, 1),
                "edad": 39,
                "quinquenio": 40
            }
        )
        workbook = load_workbook(BytesIO(contenido), data_only=False)

        self.assertEqual(
            workbook.sheetnames,
            ["Facturas", "Resumen Refinanciamiento"]
        )
        self.assertTrue(workbook["Facturas"]["A2"].value)
        self.assertEqual(workbook["Facturas"]["B2"].value, "12546")
        self.assertEqual(workbook["Facturas"]["D2"].value, 42117.20)
        self.assertEqual(
            workbook["Facturas"]["K2"].number_format,
            '0"%"'
        )
        self.assertEqual(
            workbook["Resumen Refinanciamiento"]["B3"].value,
            23
        )
        self.assertEqual(
            workbook["Resumen Refinanciamiento"]["B4"].value,
            "Juan Manuel"
        )
        self.assertEqual(
            workbook["Resumen Refinanciamiento"]["B11"].value,
            2506.54
        )
        self.assertEqual(
            workbook["Resumen Refinanciamiento"]["B24"].value,
            121282.17
        )
        self.assertEqual(
            workbook["Resumen Refinanciamiento"]["B25"].value,
            180470.88
        )
        self.assertTrue(
            workbook["Resumen Refinanciamiento"]["A24"].font.bold
        )
        self.assertEqual(
            workbook["Resumen Refinanciamiento"]["A24"].fill.fgColor.rgb,
            "00FFF2CC"
        )
        self.assertTrue(
            workbook["Resumen Refinanciamiento"]["A26"].font.bold
        )

    def test_construye_y_guarda_en_carpeta_por_semana_vendedor_cliente(self):
        resumen = calcular_resumen_refinanciamiento(self.facturas, 0)
        contenido = generar_excel_refinanciamiento(
            facturas=self.facturas,
            resumen=resumen,
            datos_cliente={
                "semana": 23,
                "vendedor": "Juan Manuel",
                "fecha": date(2026, 6, 9),
                "cliente": "JUAN CARLOS NAVA CASTRO",
                "rfc_nac": "NACJ860901UT6",
                "fecha_nacimiento": date(1986, 9, 1),
                "edad": 39,
                "quinquenio": 40
            }
        )

        with tempfile.TemporaryDirectory() as carpeta_temporal:
            ruta = guardar_excel_refinanciamiento(
                contenido=contenido,
                base_dir=carpeta_temporal,
                semana=23,
                vendedor="Juan Manuel",
                cliente="JUAN CARLOS NAVA CASTRO"
            )
            esperada = (
                Path(carpeta_temporal)
                / "SEM_23"
                / "Juan Manuel"
                / "JUAN CARLOS NAVA CASTRO"
                / "refinanciamiento.xlsx"
            )

            self.assertEqual(ruta, esperada)
            self.assertTrue(ruta.exists())

    def test_sanitiza_nombres_y_genera_nombre_de_archivo(self):
        self.assertEqual(
            slug_folder_name(' Juan / Carlos: Nava  '),
            "Juan Carlos Nava"
        )
        self.assertEqual(
            str(construir_carpeta_refinanciamiento(
                "salidas",
                1,
                "",
                ""
            )),
            str(
                Path("salidas")
                / "SEM_01"
                / "VENDEDOR_SIN_SELECCIONAR"
                / "CLIENTE_SIN_NOMBRE"
            )
        )
        self.assertEqual(
            nombre_archivo_refinanciamiento("", 9),
            "refinanciamiento.xlsx"
        )

    def test_prepara_facturas_bd_y_selecciona_solo_aptas(self):
        resultado = preparar_facturas_desde_bd(pd.DataFrame([
            {
                "venta_id": 1,
                "fact": "APTA",
                "vta": 1000,
                "pagado_db": 400,
                "saldo": 600
            },
            {
                "venta_id": 2,
                "fact": "NO APTA",
                "vta": 1000,
                "pagado_db": 399,
                "saldo": 601
            }
        ]))

        self.assertTrue(resultado.loc[0, "INCLUIR"])
        self.assertFalse(resultado.loc[1, "INCLUIR"])
        self.assertEqual(resultado.loc[0, "PAGADO"], 400)
        self.assertEqual(resultado.loc[0, "SALDO"], 600)
        self.assertEqual(resultado.loc[0, "ABONO"], 13.89)
        self.assertEqual(resultado.loc[0, "ESTATUS"], "APTA")
        self.assertEqual(resultado.loc[1, "ESTATUS"], "NO APTA")

    def test_ejemplo_excel_desde_pagado_db(self):
        facturas_bd = pd.DataFrame([
            {"fact": "12546", "vta": 58564, "pagado_db": 42117.20},
            {"fact": "12547", "vta": 28693.43, "pagado_db": 20324.52},
            {"fact": "12622", "vta": 29966.96, "pagado_db": 19978.08},
            {"fact": "12801", "vta": 22251.39, "pagado_db": 12671.05},
            {"fact": "12719", "vta": 40994.80, "pagado_db": 26191.02}
        ])

        resultado = preparar_facturas_desde_bd(facturas_bd)
        resumen = calcular_resumen_refinanciamiento(resultado, 0)

        self.assertTrue(resultado["INCLUIR"].all())
        porcentajes = dict(zip(
            resultado["FACT"].astype(str),
            resultado["PORCENTAJE PAGADO"].round()
        ))
        self.assertEqual(
            porcentajes,
            {
                "12546": 72,
                "12547": 71,
                "12622": 67,
                "12801": 57,
                "12719": 64
            }
        )
        self.assertEqual(resumen["total_vta"], 180470.58)
        self.assertEqual(resumen["total_pagado"], 121281.87)
        self.assertEqual(resumen["total_saldo_pendiente"], 59188.71)
        self.assertEqual(resumen["abono_ref"], 2506.54)
        self.assertEqual(
            resumen["simulacion"][72]["VENTA POSIBLE"],
            121282.17
        )

    def test_incluir_persiste_por_sale_id_sin_depender_del_indice(self):
        facturas = preparar_facturas_desde_bd(pd.DataFrame([
            {
                "venta_id": 501,
                "fact": "APTA",
                "vta": 1000,
                "pagado_db": 500
            },
            {
                "venta_id": 502,
                "fact": "NO APTA",
                "vta": 1000,
                "pagado_db": 300
            }
        ]))
        selecciones = {}
        aplicar_inclusiones_por_sale_id(facturas, selecciones)

        facturas_editadas = facturas.copy()
        facturas_editadas.loc[
            facturas_editadas["VENTA_ID"].eq(501),
            "INCLUIR"
        ] = False
        facturas_editadas.loc[
            facturas_editadas["VENTA_ID"].eq(502),
            "INCLUIR"
        ] = True
        actualizar_inclusiones_por_sale_id(
            facturas_editadas,
            selecciones
        )

        recalculadas = calcular_facturas_refinanciamiento(
            facturas.sample(frac=1, random_state=7).reset_index(drop=True)
        )
        restauradas = aplicar_inclusiones_por_sale_id(
            recalculadas,
            selecciones
        ).set_index("VENTA_ID")

        self.assertFalse(restauradas.loc[501, "INCLUIR"])
        self.assertTrue(restauradas.loc[502, "INCLUIR"])
        resumen = calcular_resumen_refinanciamiento(restauradas, 0)
        self.assertEqual(resumen["total_vta"], 1000)

    def test_saldo_mayor_a_vta_se_marca_para_revision(self):
        resultado = calcular_facturas_refinanciamiento(pd.DataFrame([{
            "FACT": "ANOMALA",
            "VTA": 1000,
            "SALDO": 1250,
            "QNAS TOMADAS A CUENTA": 0,
            "ABONO": 100,
            "EN COBRO": ""
        }]))

        self.assertEqual(resultado.loc[0, "PAGADO"], -250)
        self.assertEqual(resultado.loc[0, "PORCENTAJE PAGADO"], 0)
        self.assertEqual(resultado.loc[0, "ESTATUS"], "REVISAR SALDO")
        self.assertEqual(resultado.loc[0, "MOTIVO"], "SALDO mayor que VTA")
        self.assertFalse(resultado.loc[0, "INCLUIR"])

    def test_resumen_solo_considera_facturas_incluidas(self):
        facturas = self.facturas.head(2).copy()
        facturas["INCLUIR"] = [True, False]

        resumen = calcular_resumen_refinanciamiento(facturas, 0)

        self.assertEqual(resumen["total_vta"], 58564)
        self.assertEqual(resumen["total_pagado"], 42117.20)
        self.assertEqual(resumen["abono_ref"], 813.39)

    def test_genera_pdf_json_y_guarda_paquete(self):
        facturas = calcular_facturas_refinanciamiento(
            self.facturas.head(1)
        )
        resumen = calcular_resumen_refinanciamiento(facturas, 0)
        datos_cliente = {
            "fecha": date(2026, 6, 10),
            "semana": 24,
            "vendedor": "Juan Manuel",
            "cliente": "CLIENTE PRUEBA",
            "rfc_nac": "J860901UT6",
            "fecha_nacimiento": date(1986, 9, 1),
            "edad": 39,
            "quinquenio": 40
        }
        pdf = generar_pdf_refinanciamiento(
            facturas,
            resumen,
            datos_cliente
        )
        json_data = generar_json_refinanciamiento(
            facturas,
            resumen,
            datos_cliente
        )

        self.assertTrue(pdf.startswith(b"%PDF"))
        self.assertEqual(
            json.loads(json_data)["datos_cliente"]["cliente"],
            "CLIENTE PRUEBA"
        )

        with tempfile.TemporaryDirectory() as carpeta_temporal:
            rutas = guardar_archivos_refinanciamiento(
                {"xlsx": b"xlsx", "pdf": pdf, "json": json_data},
                carpeta_temporal,
                24,
                "Juan Manuel",
                "CLIENTE PRUEBA"
            )
            self.assertEqual(
                set(rutas),
                {"xlsx", "pdf", "json"}
            )
            self.assertTrue(all(ruta.exists() for ruta in rutas.values()))


if __name__ == "__main__":
    unittest.main()
