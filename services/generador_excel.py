from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


def limpiar_nombre_archivo(texto: str) -> str:
    texto = texto.upper().strip()
    caracteres_invalidos = ['\\', '/', ':', '*', '?', '"', '<', '>', '|']

    for caracter in caracteres_invalidos:
        texto = texto.replace(caracter, "")

    texto = texto.replace(" ", "_")
    return texto


def escribir_celda(ws, celda: str, valor):
    """
    Escribe en una celda normal o combinada.
    Si la celda pertenece a un rango combinado, escribe en la celda principal.
    """

    for rango in ws.merged_cells.ranges:
        if celda in rango:
            celda_principal = rango.coord.split(":")[0]
            ws[celda_principal] = valor
            return

    ws[celda] = valor


def detalle_cuentas(
    cuentas_terminadas: list[dict],
    sumar_a_liquidez: bool
) -> str:
    detalles = []

    for cuenta in cuentas_terminadas:
        if bool(cuenta.get("sumar_a_liquidez", False)) != sumar_a_liquidez:
            continue

        qna = str(cuenta.get("qna_termina", "")).strip() or "Sin QNA"
        monto = float(cuenta.get("saldo_liberado", 0) or 0)
        observacion = str(cuenta.get("observacion", "")).strip()
        detalle = f"{qna}: ${monto:,.2f}"

        if observacion:
            detalle += f" | {observacion}"

        detalles.append(detalle)

    return "\n".join(detalles)


def agregar_hoja_cuentas_terminadas(
    wb,
    datos: dict,
    revision: dict,
    promotor: str,
    qna: str
):
    nombre_hoja = "CUENTAS_TERMINADAS"

    if nombre_hoja in wb.sheetnames:
        del wb[nombre_hoja]

    ws = wb.create_sheet(nombre_hoja)
    encabezados = [
        "fecha_revision",
        "cliente",
        "rfc",
        "vendedor",
        "qna_revision",
        "qna_termina",
        "saldo_liberado",
        "sumar_a_liquidez",
        "observacion"
    ]
    ws.append(encabezados)

    fecha_revision = datetime.now().strftime("%d/%m/%Y")

    for cuenta in revision.get("cuentas_terminadas", []):
        ws.append([
            fecha_revision,
            datos["nombre"],
            datos["rfc"],
            promotor,
            qna,
            cuenta.get("qna_termina", ""),
            float(cuenta.get("saldo_liberado", 0) or 0),
            "Sí" if cuenta.get("sumar_a_liquidez", False) else "No",
            cuenta.get("observacion", "")
        ])

    color_encabezado = "1F4E78"
    for celda in ws[1]:
        celda.fill = PatternFill("solid", fgColor=color_encabezado)
        celda.font = Font(color="FFFFFF", bold=True)
        celda.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{max(ws.max_row, 1)}"
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 20
    ws.column_dimensions["I"].width = 45

    for celda in ws["G"][1:]:
        celda.number_format = '$#,##0.00;-$#,##0.00;$-'


def generar_excel_revision(
    datos: dict,
    revision: dict,
    mensaje_vendedor: str,
    promotor: str,
    qna: str = "09-2026",
    ruta_template: str = "templates/plantilla_revision_talon.xlsx",
    carpeta_salida: str = "output"
) -> str:
    Path(carpeta_salida).mkdir(exist_ok=True)

    wb = load_workbook(ruta_template)
    ws = wb["Hoja1"]

    codigos = revision["codigos_revision"]

    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass

    # =========================
    # ENCABEZADO
    # =========================

    escribir_celda(ws, "B3", datos["rfc"])
    escribir_celda(ws, "D3", datos["nombre"])

    escribir_celda(ws, "C4", qna)
    escribir_celda(ws, "E4", datos.get("fecha_pago", ""))

    # =========================
    # INGRESOS
    # Mapeo correcto según tu plantilla.
    # =========================

    escribir_celda(ws, "B6", codigos["E4"])
    escribir_celda(ws, "B7", codigos["E3"])
    escribir_celda(ws, "B8", codigos["Q"])
    escribir_celda(ws, "B9", codigos["CP"])
    escribir_celda(ws, "B10", codigos["7"])
    escribir_celda(ws, "B11", codigos["CT"])
    escribir_celda(ws, "B12", codigos["7B"])
    escribir_celda(ws, "B13", codigos["E9"])
    escribir_celda(ws, "B14", codigos["SG"])
    escribir_celda(ws, "B15", codigos["O1"])

    # Total ingresos
    escribir_celda(ws, "C6", revision["ingresos"])

    # =========================
    # DEDUCCIONES
    # =========================

    escribir_celda(ws, "E6", revision["descuentos"])
    escribir_celda(ws, "E7", revision["saldo_100"])
    escribir_celda(ws, "E8", revision["total_para_venta_70"])
    escribir_celda(ws, "E10", revision["saldo_70"])
    escribir_celda(ws, "D12", revision["saldo_70"])
    escribir_celda(ws, "D15", revision["saldo_100"])

    # DC TALON y DC SISTEMA
    escribir_celda(ws, "B16", codigos["DC"])
    escribir_celda(ws, "D16", codigos["DC"])

    # =========================
    # NO ESCRIBIR VENDEDOR EN A18
    # A18:A28 está combinado verticalmente y por eso se ve letra por letra.
    # Lo dejamos limpio.
    # =========================

    escribir_celda(ws, "A18", "")

    # =========================
    # RECUPERACIÓN / VENTA
    # Por ahora no estamos llenando folios de venta.
    # =========================

    escribir_celda(ws, "C20", 0)
    escribir_celda(ws, "C21", 0)
    escribir_celda(ws, "C22", 0)
    escribir_celda(ws, "C23", 0)

    escribir_celda(ws, "E20", 0)
    escribir_celda(ws, "E21", 0)
    escribir_celda(ws, "E22", 0)
    escribir_celda(ws, "E23", 0)

    escribir_celda(ws, "B25", 0)
    escribir_celda(ws, "D25", 0)

    escribir_celda(ws, "B28", revision["saldo_70"])
    escribir_celda(ws, "D28", revision["saldo_100"])

    # =========================
    # REVISIÓN DEL CLIENTE
    # =========================

    escribir_celda(ws, "B31", promotor)
    escribir_celda(ws, "B32", datos["nombre"])
    escribir_celda(ws, "B33", datos["rfc"])
    escribir_celda(ws, "B34", revision["liquidez_final"])
    escribir_celda(ws, "B35", qna)

    escribir_celda(ws, "B36", mensaje_vendedor)
    escribir_celda(ws, "B40", datetime.now().strftime("%d/%m/%Y"))

    cuentas_terminadas = revision.get("cuentas_terminadas", [])
    detalle_liberadas = detalle_cuentas(cuentas_terminadas, True)
    detalle_observadas = detalle_cuentas(cuentas_terminadas, False)

    ws.merge_cells("A41:C41")
    ws.merge_cells("D41:E41")
    ws.merge_cells("A42:C42")
    ws.merge_cells("D42:E42")
    ws.merge_cells("A43:C43")
    ws.merge_cells("D43:E43")
    ws.merge_cells("A44:C44")
    ws.merge_cells("D44:E44")
    ws.merge_cells("A45:C45")
    ws.merge_cells("D45:E45")
    ws.merge_cells("A46:C46")
    ws.merge_cells("D46:E46")
    ws.merge_cells("A47:C47")
    ws.merge_cells("D47:E47")
    ws.merge_cells("A48:E48")
    ws.merge_cells("A49:E51")
    ws.merge_cells("A52:E52")
    ws.merge_cells("A53:E55")

    escribir_celda(ws, "A41", "LIQUIDEZ DEL TALÓN")
    escribir_celda(ws, "D41", revision.get("liquidez_talon", revision["saldo_100"]))
    escribir_celda(ws, "A42", "APOYO ADICIONAL")
    escribir_celda(ws, "D42", revision.get("abono_extra", 0))
    escribir_celda(ws, "A43", "TOTAL SALDO LIBERADO")
    escribir_celda(ws, "D43", revision.get("total_saldo_liberado", 0))
    escribir_celda(ws, "A44", "TOTAL SOLO OBSERVADO")
    escribir_celda(ws, "D44", revision.get("total_solo_observado", 0))
    escribir_celda(ws, "A45", "MONTO PROGRAMADO")
    escribir_celda(ws, "D45", revision.get("programado", 0))
    escribir_celda(ws, "A46", "LIQUIDEZ ANTES LIBERACIÓN")
    escribir_celda(ws, "D46", revision.get("liquidez_antes_liberacion", revision["liquidez_final"]))
    escribir_celda(ws, "A47", "LIQUIDEZ FINAL")
    escribir_celda(ws, "D47", revision["liquidez_final"])
    escribir_celda(ws, "A48", "DETALLE CUENTAS LIBERADAS")
    escribir_celda(ws, "A49", detalle_liberadas or "Sin cuentas sumadas a liquidez")
    escribir_celda(ws, "A52", "DETALLE CUENTAS OBSERVADAS")
    escribir_celda(ws, "A53", detalle_observadas or "Sin cuentas solo observadas")

    for celda in ["A41", "A42", "A43", "A44", "A45", "A46", "A47", "A48", "A52"]:
        ws[celda].font = Font(bold=True)

    ws["A49"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A53"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[49].height = max(30, 15 * max(len(detalle_liberadas.splitlines()), 1))
    ws.row_dimensions[53].height = max(30, 15 * max(len(detalle_observadas.splitlines()), 1))
    ws.print_area = "A1:E55"

    # =========================
    # FORMATO DE MONEDA
    # =========================

    celdas_moneda = [
        "B6", "B7", "B8", "B9", "B10", "B11", "B12", "B13", "B14", "B15",
        "C6",
        "E6", "E7", "E8", "E10", "D12", "D15",
        "B16", "D16",
        "B25", "D25", "B28", "D28",
        "B34", "D41", "D42", "D43", "D44", "D45", "D46", "D47"
    ]

    for celda in celdas_moneda:
        escribir_celda(ws, celda, ws[celda].value)
        ws[celda].number_format = '$#,##0.00;-$#,##0.00;$-'

    agregar_hoja_cuentas_terminadas(
        wb=wb,
        datos=datos,
        revision=revision,
        promotor=promotor,
        qna=qna
    )

    # =========================
    # GUARDAR ARCHIVO
    # =========================

    nombre_cliente = limpiar_nombre_archivo(datos["nombre"])
    nombre_promotor = limpiar_nombre_archivo(promotor)

    nombre_archivo = f"REVISION_{nombre_cliente}_{nombre_promotor}.xlsx"
    ruta_salida = Path(carpeta_salida) / nombre_archivo

    wb.save(ruta_salida)

    return str(ruta_salida)
